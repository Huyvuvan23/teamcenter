import os
import re
import time
import sys
sys.coinit_flags = 2  # COINIT_APARTMENTTHREADED
import pyperclip
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl
import xlwings as xw
from pywinauto.application import Application
from pywinauto import keyboard as kb
import win32api, win32con
import pyautogui as pag
import keyboard
import signal
import shutil
import psutil
import threading
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
from PyQt5 import QtCore
from PyQt5.QtCore import pyqtSignal
from datetime import datetime, timedelta
import json
import ctypes
import xml.etree.ElementTree as ET
from ui import MainWindow 

# Redirect stdout to both terminal and file
class Logger:
    def __init__(self, filename):
        self.terminal = sys.stdout
        self.log = open(filename, "w")

    def write(self, message):
        self.terminal.write(message)  # Print to terminal
        self.terminal.flush()         # Force terminal update
        self.log.write(message)      # Write to file
        self.log.flush()             # Force file write (no buffer delay)

    def flush(self):
        pass  # Needed for Python 3 compatibility

class TeamcenterDownloader(QtCore.QObject):
    update_status = pyqtSignal(str, str)  # process_label, noti_label
    enable_download_button = pyqtSignal(bool)
    show_message = pyqtSignal(str, str, str)  # title, text, type

    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.ui = main_window.ui
        self.data_file = "form_data.json"
        self.data = {}
        self.stop_flag = False
        if os.path.exists(self.data_file):
            with open(self.data_file, 'r') as f:
                self.data = json.load(f)

        self.update_status.connect(self._update_status)
        self.enable_download_button.connect(self.ui.download_button.setEnabled)
        self.show_message.connect(self._show_message)

        self.load_data()
        self.connect_ui()

        # Set window close handler for PyQt
        self.main_window.closeEvent = self._on_close_event

    def _update_status(self, process_text, noti_text):
        if process_text:
            self.ui.process_label.setText(process_text)
        
        if noti_text:
            self.ui.noti_label.setText(noti_text)

    def _show_message(self, title, text, type_):
        msg_box = QMessageBox(self.main_window)
        if type_ == "error":
            msg_box.setIcon(QMessageBox.Critical)
        else:
            msg_box.setIcon(QMessageBox.Information)
        msg_box.setWindowTitle(title)
        msg_box.setText(text)
        msg_box.setStyleSheet("QLabel{-family:'Segoe UI';} QPushButton{-family:'Segoe UI';}")
        msg_box.exec_()

    def connect_ui(self):
        self.ui.input_file_button_1.clicked.connect(lambda: self.select_input_file(self.ui.input_file_entry_1, "MAP" if self.ui.input_file_label_1.text().lower().startswith("map") else ""))
        self.ui.input_file_button_2.clicked.connect(lambda: self.select_input_file(self.ui.input_file_entry_2, "Connector IF"))
        self.ui.output_button.clicked.connect(lambda: self.select_output_folder(self.ui.output_folder_entry))
        self.ui.download_button.clicked.connect(self.download)
        self.ui.stop_button.clicked.connect(self.stop_task)

    def save_data(self):
        data = {
            "input_file_1": self.ui.input_file_entry_1.text(),
            "input_file_2": self.ui.input_file_entry_2.text(),
            "output_folder": self.ui.output_folder_entry.text(),
            "coliteam": self.ui.item_column_entry.text(),
            "colrevision": self.ui.rev_column_entry.text(),
            "colnamefolder": self.ui.folder_column_entry.text(),
            "var": self.ui.file_mode_toggle.isChecked(),
        }
        with open(self.data_file, 'w') as f:
            json.dump(data, f)

    def _on_close_event(self, event):
        self.save_data()
        event.accept()

    def load_data(self):
        if os.path.exists(self.data_file):
            with open(self.data_file, 'r') as f:
                self.data = json.load(f)
        else:
            self.data = {}
        # Set UI values if data exists
        d = self.data
        self._set_text(self.ui.input_file_entry_1, d.get("input_file_1", ""))
        self._set_text(self.ui.input_file_entry_2, d.get("input_file_2", ""))
        self._set_text(self.ui.output_folder_entry, d.get("output_folder", ""))
        self._set_text(self.ui.item_column_entry, d.get("coliteam", ""))
        self._set_text(self.ui.rev_column_entry, d.get("colrevision", ""))
        self._set_text(self.ui.folder_column_entry, d.get("colnamefolder", ""))
        self.ui.file_mode_toggle.setChecked(d.get("var", False))
    
    def _set_text(self, widget, value):
        widget.setText(value)

    def select_input_file(self, entry_widget, text):
        file_path, _ = QFileDialog.getOpenFileName(self.main_window, f"Select File {text}", None, "Excel Files (*.xls*);;All Files (*)")

        if file_path:
            file_path = file_path.replace("/", "\\")
            entry_widget.setText(file_path)

    def select_output_folder(self, entry_widget):
        folder_path = QFileDialog.getExistingDirectory(self.main_window, "Select Folder")
        if folder_path:
            folder_path = folder_path.replace("/", "\\")
            entry_widget.setText(folder_path)

    def waiting_progress(self, teamcenter_window):
        time_start = time.time()
        while True:
            if self.stop_flag: return
            if time.time() - time_start >= 5*60:
                self.update_status.emit(
                    "",
                    "<html><head/><body><p align='center' style='Stopped!</p></body></html>"
                )
                print("Out of waiting time")
                sys.exit()
            if teamcenter_window.child_window(control_type="Text", title="No operations to display at this time.").exists():
                return False

    def reset(self, teamcenter_window):
        if self.stop_flag: return 
        teamcenter_window.set_focus()
        kb.send_keys("%WOM")
        self.waiting_progress(teamcenter_window)
        kb.send_keys("%WRY")
        self.waiting_progress(teamcenter_window)
        try:
            teamcenter_window.child_window(title="Close All", control_type="MenuItem").select()
        except:
            pass

    def preparing(self, teamcenter_window, type):
        if self.stop_flag: return
        teamcenter_window.child_window(control_type="SplitButton", title="Select a Search").invoke()
        self.waiting_progress(teamcenter_window)

        if type == 'excel':
            kb.send_keys("I{ENTER}")
        if type == 'zip':
            kb.send_keys("R{ENTER}")
        if type == 'nx':    
            kb.send_keys("SSS{ENTER}")

        self.waiting_progress(teamcenter_window)
        teamcenter_window.child_window(control_type="Button", title="Clear all search fields").invoke()

    def get_data_from_simple_file(self, file_link, colnamefolder, coliteam, colrevision):
        data = []
        try:
            wb = load_workbook(filename=file_link)
            sheet = wb.active
            maxrow = sheet.max_row
                
            for i in range(2, maxrow + 1):  # Start from row 2, assuming row 1 is header
                item = sheet[f'{coliteam}{i}'].value
                revision = sheet[f'{colrevision}{i}'].value
                folder_name = sheet[f'{colnamefolder}{i}'].value
                header_values = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                if header_values == ["Unit", "Shape No", "Shape Revision", "Data Note", "Rev Drawing", "NX PDF"]:
                    col_idx = openpyxl.utils.column_index_from_string(colrevision)
                    stt_datanote = sheet.cell(row=i, column=col_idx + 1).value
                    stt_rev = sheet.cell(row=i, column=col_idx + 2).value
                    stt_pdf = sheet.cell(row=i, column=col_idx + 3).value
                else:
                    stt_datanote = None
                    stt_rev = None
                    stt_pdf = None
                # Only add if all fields are non-empty strings
                if all(isinstance(val, str) and val.strip() for val in [item, revision, folder_name]):
                    tup = (folder_name.strip(), item.strip(), revision.strip(), stt_datanote, stt_rev, stt_pdf)
                    if tup not in data:
                        data.append(tup)
        except Exception as e:
            print(f"Error reading input file: {e}")
        return data

    def get_data_from_map_file(self, link_MAP_file, link_connector_infor_file):
        def set_up_ws(ws):
            try:
                ws.api.Rows.Hidden = False
                ws.api.Columns.Hidden = False
                ws.api.AutoFilterMode = False
            except Exception:
                pass

        def get_last_row(ws, col):
            return ws.range(col + str(ws.cells.last_cell.row)).end('up').row

        def find_cell(ws, text_to_find):
            unit_name_cell_ws_if = None
            try:
                found = ws.api.Cells.Find(
                    What = text_to_find,
                    LookIn=-4163,  # xlValues
                    LookAt=1,      # xlWhole
                    SearchOrder=1, # xlByRows
                    SearchDirection=1, # xlNext
                    MatchCase=False
                )
                if found is not None:
                    unit_name_cell_ws_if = ws.range((found.Row, found.Column))
            except Exception:
                pass
            return unit_name_cell_ws_if
        
        app = xw.App(visible=False)
        wb_if = None
        wb_map = None

        try:
            wb_if = app.books.open(link_connector_infor_file)
            ws_if = wb_if.sheets("コネクタIFの作成管理")
            set_up_ws(ws_if)

            unit_name_cell_ws_if = find_cell(ws_if, "unit name")
            
            # Get the last row and all unit names from the "unit name" column in the IF sheet
            if unit_name_cell_ws_if is None:
                raise ValueError("'unit name' cell not found in コネクタIFの作成管理 sheet.")
            unit_name_cell_col_if = get_column_letter(unit_name_cell_ws_if.column)
            unit_name_cell_row_if = unit_name_cell_ws_if.row
            last_row_if = get_last_row(ws_if, unit_name_cell_col_if)
            unitnames_if = ws_if.range(f"{unit_name_cell_col_if}{unit_name_cell_row_if + 3}:{unit_name_cell_col_if}{last_row_if}").value
            unitnames_if = sorted(set(x for x in unitnames_if if x is not None))

            wb_map = app.books.open(link_MAP_file)
            ws_map = wb_map.sheets("MAP")
            ws_dwg = wb_map.sheets["DWG"]
            set_up_ws(ws_map)
            set_up_ws(ws_dwg)
            
            unit_name_cell_map = find_cell(ws_map, "unit name")
            part_number_cell_map = find_cell(ws_map, "Part#(10)")

            # Use the column letter for "unit name" dynamically
            unit_name_cell_col_map = get_column_letter(unit_name_cell_map.column)
            unit_name_cell_row_map = unit_name_cell_map.row
            last_row_map = get_last_row(ws_map, unit_name_cell_col_map)
            unitname_map = ws_map.range(f"{unit_name_cell_col_map}{unit_name_cell_row_map + 1}:{unit_name_cell_col_map}{last_row_map}").value

            part_number_cell_col_map = get_column_letter(part_number_cell_map.column)
            part_number_map = ws_map.range(f"{part_number_cell_col_map}{unit_name_cell_row_map + 1}:{part_number_cell_col_map}{last_row_map}").value

            no_cell_dwg = find_cell(ws_dwg, "no")
            no_cell_col_dwg = get_column_letter(no_cell_dwg.column)
            no_cell_row_dwg = no_cell_dwg.row
            last_row_dwg = get_last_row(ws_dwg, no_cell_col_dwg)
            
            # Get the columns by checking the header row
            headers = ws_dwg.range(f"A{no_cell_row_dwg}:XFD{no_cell_row_dwg}").value
            try:
                part_number_column = headers.index("PART #") + 1
                shape_no_column = headers.index("Drawing/Shape No") + 1
                shape_revision_column = headers.index("Drawing/Shape Revision") + 1
            except ValueError as e:
                raise ValueError("One or more necessary column headers not found.") from e
            
            data_start = no_cell_row_dwg + 1
            list_part_number_dwg = ws_dwg.range(ws_dwg.cells(data_start, part_number_column),
                                                ws_dwg.cells(last_row_dwg, part_number_column)).value
            list_shape_no_dwg = ws_dwg.range(ws_dwg.cells(data_start, shape_no_column),
                                            ws_dwg.cells(last_row_dwg, shape_no_column)).value
            list_shape_revision_dwg = ws_dwg.range(ws_dwg.cells(data_start, shape_revision_column),
                                                ws_dwg.cells(last_row_dwg, shape_revision_column)).value
            
            list_shape_merged_dwg = [
                f"{no_val},{rev_val}" if no_val is not None and rev_val is not None else ""
                for no_val, rev_val in zip(list_shape_no_dwg, list_shape_revision_dwg)
            ]
            
            # Group part numbers based on unit names from the "IF" sheet.
            part_numbers_by_unit = []
            for unit in unitnames_if:
                indices = [i for i, nm in enumerate(unitname_map) if nm == unit]
                values = sorted(set(part_number_map[i] for i in indices if part_number_map[i] is not None))
                part_numbers_by_unit.append(values)
            
            # Get the shapes for each unit
            shapes_by_unit = []
            for part_numbers in part_numbers_by_unit:
                unit_shapes = []
                for part in part_numbers:
                    indices = [i for i, pn in enumerate(list_part_number_dwg) if pn == part]
                    unit_shapes.extend(list_shape_merged_dwg[i] for i in indices)
                shapes_by_unit.append(sorted(set(unit_shapes)))
            
            all_data = []
            for idx, unit_shapes in enumerate(shapes_by_unit):
                is_ok = False
                for item in unit_shapes:
                    if item:
                        is_ok = True
                        shape_parts = item.split(',')
                        if len(shape_parts) == 2:
                            all_data.append((unitnames_if[idx], shape_parts[0], shape_parts[1], None, None, None))
                if not is_ok:
                    all_data.append((unitnames_if[idx], "NONE", "NONE", None, None, None))
            
            return all_data
        finally:
            if wb_if:
                wb_if.close()
            if wb_map:
                wb_map.close()
            app.quit()

    def get_latest_excel_zip_file(self):
        folder_path = rf'C:\Temp'
        folders = [f for f in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, f))]
        if not folders:
            return None
        
        full_folder_paths = [os.path.join(folder_path, f) for f in folders]
        latest_folder = max(full_folder_paths, key=os.path.getmtime)
        
        excel_files = [f for f in os.listdir(latest_folder) if (f.endswith('.xlsm') or f.endswith('.xls') or f.endswith('.zip') or f.endswith('.xlsx'))  and not f.startswith('~$')]
        if not excel_files:
            return None
        
        full_file_paths = [os.path.join(latest_folder, f) for f in excel_files]
        latest_file = max(full_file_paths, key=os.path.getmtime)
        
        return latest_file

    def copy_latest_excel_zip_file(self, new_folder_moi, tenmoi):
        latest_file = self.get_latest_excel_zip_file()
        if latest_file:
            destination_folder = new_folder_moi
            destination_path = os.path.join(destination_folder, os.path.basename(latest_file))
            
            base, extension = os.path.splitext(destination_path)
            destination_path = f"{base}_{tenmoi}{extension}"
            counter = 1
            original_base = base
            while os.path.exists(destination_path):
                destination_path = f"{original_base}_{tenmoi}_{counter}{extension}"
                counter += 1
            
            shutil.copy2(latest_file, destination_path)
            print(f"Latest file '{os.path.basename(latest_file)}' copied to '{destination_folder}'.")
        else:
            print("No file found.")

    def kill_new_excel_processes(self):
        threshold_time = datetime.now() - timedelta(minutes=2)
        for proc in psutil.process_iter(['pid', 'name', 'create_time']):
            if proc.info['name'] == 'EXCEL.EXE':
                process_create_time = datetime.fromtimestamp(proc.info['create_time'])
                if process_create_time > threshold_time:
                    os.kill(proc.info['pid'], signal.SIGTERM)

    def kill_new_7zip_processes(self):
        threshold_time = datetime.now() - timedelta(minutes=2)
        for proc in psutil.process_iter(['pid', 'name', 'create_time']):
            if proc.info['name'] =='7zFM.exe':
                process_create_time = datetime.fromtimestamp(proc.info['create_time'])
                if process_create_time > threshold_time:
                    os.kill(proc.info['pid'], signal.SIGTERM)

    def confirm_window_openned(self, teamcenter_app, teamcenter_window):
        star_time = time.time()
        confirm_window = teamcenter_window.child_window(class_name="SunAwtDialog")
        while True:
            if self.stop_flag: return
            windows = teamcenter_app.windows()
            for window in windows:
                try:
                    if window != teamcenter_window:
                        window.set_focus()
                        time.sleep(1)
                        kb.send_keys("{ENTER}")
                    else:
                        if confirm_window . exists():
                            confirm_window . child_window(title="Close", control_type="Button").invoke()
                            return False
                except:
                    pass
            if time.time() - star_time >= 300:
                print("Out of waiting time")
                sys.exit()
            
            else:
                if teamcenter_window.child_window(control_type="Text", title="No operations to display at this time.").exists():
                    return True
                else:
                    continue

    def create_folder(self, outputfolder, folder_name):
        # Replace all characters not allowed in folder names with "_"
        sanitized_folder_name = re.sub(r'[<>:"/\\|?*]', '_', str(folder_name))
        folder_path = os.path.join(outputfolder, sanitized_folder_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        return folder_path

    def set_search_fields(self, search_window, item_id, revision, file_type):
        
        if file_type == 'excel':
            # For Data Note (Excel), set Item ID and Revision fields
            search_window.child_window(control_type="Edit", title="Item ID:").set_text(item_id)
            search_window.child_window(control_type="Edit", title="Revision:").set_text(revision)
        elif file_type == 'zip':
            # For Ref Drawing (ZIP), set Shape Number and Shape Change Number fields
            search_window.child_window(control_type="Edit", title="Shape Number:").set_text(item_id)
            search_window.child_window(control_type="Edit", title="Shape Change Number:").set_text(revision)
        elif file_type == 'nx':
            # For NX PDF, handle ShapeNumber and ShapeName fields
            shapenumber = search_window.child_window(control_type="Edit", title="ShapeNumber:")
            shapename = search_window.child_window(control_type="Edit", title="ShapeName:")
            btn_more = search_window.child_window(title="More...>>>", control_type="Button")
            i = 0
            # If item_id ends with '*', use ShapeNumber, else use ShapeName
            if item_id.endswith("*"):
                # Ensure ShapeNumber field is visible
                while not shapenumber.exists() and i < 10:
                    btn_more.click()
                    time.sleep(0.5)
                    i += 1
                    shapename.set_text("")
                if not shapenumber.exists() : self.preparing(search_window.parent(), "nx")
                shapenumber.set_text(item_id)
            else:
                # Ensure ShapeName field is visible
                while not shapename.exists() and i < 10:
                    btn_more.click()
                    time.sleep(0.5)
                    i += 1
                    shapenumber.set_text("")
                if not shapename.exists() : self.preparing(search_window.parent(), "nx")
                shapename.set_text(item_id)
            # Set ShapeChangeNumber field
            search_window.child_window(control_type="Edit", title="ShapeChangeNumber:").set_text(revision)

    def download_file(self, teamcenter_app, teamcenter_window, outputfolder, folder_name, item_id, revision, file_type):
        
        if file_type == 'excel':
            item_id_moi = item_id + "-note"
        elif file_type == 'nx':
            item_id_moi = item_id + "-shape"
        else:
            item_id_moi = item_id
        revision_moi = "0" + revision[-2:]
        
        search_window = teamcenter_window.child_window(title="Search", control_type="Tab")
        self.set_search_fields(search_window, item_id_moi, revision_moi, file_type)
        kb.send_keys("{ENTER}")
        self.waiting_progress(teamcenter_window)

        kiem_tra_item_window = teamcenter_window.child_window(title="Search Results", control_type="Tab")
        
        if file_type == 'excel':
            if kiem_tra_item_window.child_window(control_type="Pane", title="Item Revision...  - No objects found").exists():
                print("No object found for Excel")
                self.download_status = 3
                return False
        elif file_type == 'zip':
            if kiem_tra_item_window.child_window(control_type="Pane", title="Ref Drawing  - No objects found").exists():
                print("No ZIP file found")
                self.download_status = 2
                return False
        
        elif file_type == 'nx':
            status = kiem_tra_item_window.child_window(control_type="Pane", title="Shape  - No objects found")
            if status.exists():
                self.set_search_fields(search_window, item_id + "*", revision_moi, file_type)
                kb.send_keys("{ENTER}")
                self.waiting_progress(teamcenter_window)
                if status.exists():
                    print("No Shape found")
                    self.download_status = 3
                    return False

        search_result_window = kiem_tra_item_window.child_window(control_type="Pane", title="Search Results")

        def excel_and_zip():
            try:
                search_result_window.child_window(control_type="TreeItem", found_index=1).select()
            except:
                print("Cannot click TreeItem 1")
                self.download_status = 0
                return False

            self.waiting_progress(teamcenter_window)
            search_result_window.child_window(control_type="Button", title="", found_index=4).click()

            for index in range(3, 11):
                file_ez = kiem_tra_item_window.child_window(control_type="TreeItem", found_index=index)
                if file_ez.exists():
                    try:
                        file_ez.type_keys("{ENTER}")
                    except Exception as e: 
                        print(f"Have Error: {e}")
                        self.download_status = 0
                        return
                    
                    if self.confirm_window_openned(teamcenter_app, teamcenter_window):
                        time.sleep(0.5)
                        for each_folder in folder_name.split(","):
                            folder_path = self.create_folder(outputfolder, each_folder)
                            self.copy_latest_excel_zip_file(folder_path, revision_moi)

                        if file_type == 'excel':
                            self.kill_new_excel_processes()
                        else:
                            self.kill_new_7zip_processes()
                        time.sleep(0.5)
                        self.download_status = 1
                else:   
                    print(f"No file found at position {index}")
                    break
        
        def export_pdf():
            def name_pdf_cad(file_xml):
                # Parse the XML file
                tree = ET.parse(file_xml)
                root = tree.getroot()

                # Define the XML namespace (found in the root tag)
                ns = {'plm': 'http://www.plmxml.org/Schemas/PLMXMLSchema'}

                # Find the Representation element and get its 'name' attribute
                representation = root.find('.//plm:Representation', ns)
                if representation is not None:
                    name = representation.get('name')
                    return name
                else:
                    return 0
            def close_all_tab_nx():
                self.window_nx.set_focus()
                kb.send_keys('%f')
                keyboard.send('c')
                keyboard.send('l')
                time.sleep(1)
                keyboard.send('n')

            if self.first_turn:
                self.window_nx.maximize()
            kb.send_keys('^+d')
            win32api.SetCursorPos((0, 0))
            name_nx_cad = name_pdf_cad(r"C:\Temp\NX_Nav_.plmxml")
            filename = f"{name_nx_cad},{revision_moi}.pdf"

            list_name_folder = folder_name.split(",")
            for idx, each_folder in enumerate(list_name_folder):
                try:
                    
                    folder_path = self.create_folder(outputfolder, each_folder)
                    file_path = os.path.join(folder_path, filename)
                    counter = 1
                    while os.path.exists(file_path):
                        filename = f"({counter}) {name_nx_cad},{revision_moi}.pdf"
                        file_path = os.path.join(folder_path, filename)
                        counter += 1
        
                    if idx != 0:
                        try:
                            shutil.copy2(os.path.join(self.create_folder(outputfolder, list_name_folder[idx-1]), filename), file_path)
                            print(f"File '{os.path.basename(filename)}' is saved to '{each_folder}'.")
                            continue
                        except:
                            print(f"Error when sanving file '{os.path.basename(filename)}' to '{each_folder}'.")
                            return False

                    pyperclip.copy(file_path)
                    kb.send_keys('%f')
                    keyboard.send('e')
                    keyboard.send('e')
                    keyboard.send('Enter')
                    keyboard.send('d')
                    
                    if self.stop_flag: 
                        self.update_status.emit(
                            "",
                            "<html><head/><body><p align='center' style='Stopped!</p></body></html>"
                        )
                        return
                    
                    self.export_window.wait('ready', timeout=5)
                    if self.first_turn:
                        self.export_window.child_window(control_type="ComboBox",found_index=0).wrapper_object().select("File Browser")
                        kb.send_keys('{TAB}')

                    kb.send_keys('^v')
                    self.export_window.child_window(title="Current Display", control_type="ListItem").wrapper_object().invoke()
                    kb.send_keys('{DOWN}')
                    kb.send_keys('+{DOWN 10}')  # Shift+Down 10 times

                    if self.first_turn:
                        self.export_window.child_window(control_type="ComboBox",found_index=1).wrapper_object().select("Black on White")

                    self.export_window.child_window(title="OK", control_type="Button").click()
                        
                    self.window_nx.wait('ready', timeout=999)
                    self.first_turn = False
                    if os.path.exists(file_path):
                        print(f"File '{os.path.basename(filename)}' is saved to '{each_folder}'.")
                    else:
                        print(f"Error when sanving file '{os.path.basename(filename)}' to '{each_folder}'.")
                        return False
                    
                    if self.stop_flag: 
                        self.update_status.emit(
                            "",
                            "<html><head/><body><p align='center' style='Stopped!</p></body></html>"
                        )
                        return
                    
                except Exception as e:
                    print(f"Have error with NX: {e}")
                    close_all_tab_nx()
                    return False
                
            if self.total_open_NX % 5 == 0:
                close_all_tab_nx()
            return True

        def find_and_open_nx():
            def click(x, y):
                """Simulates a single mouse click at the specified (x, y) coordinates."""
                win32api.SetCursorPos((x, y))
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0)
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0)

            teamcenter_window.set_focus()
            try:
                # Select and click up to 3 TreeItems in sequence
                for idx in range(1, 4):
                    try:
                        search_result_window.child_window(control_type="TreeItem", found_index=idx).wrapper_object().select()
                        search_result_window.child_window(control_type="Button", title="", found_index=4).click()
                        self.waiting_progress(teamcenter_window)
                        
                    except Exception as e:
                        print(f"Error processing TreeItem {idx}: {e}")
                        break
                if self.stop_flag:
                    self.update_status.emit(
                        "",
                        "<html><head/><body><p align='center' style='Stopped!</p></body></html>"
                    )
                    return
                # Find all occurrences of "CAD_NX.png" on screen with error handling
                found_images = []
                num = 1
                while True:
                    image_file = rf"images\search_by_images\CAD_NX_{num}.png"
                    if not os.path.exists(image_file):
                        break
                    # Try to find all occurrences of the image on screen
                    try:
                        images_found = pag.locateAllOnScreen(image_file, confidence=0.9)
                        if images_found:
                            found_images.extend(list(images_found))
                    except Exception as e:
                        pass
                    num += 1

                coordinates = []
                for region in found_images:
                    center_x = region.left + (region.width // 2)
                    center_y = region.top + (region.height // 2)
                    coordinates.append((center_x, center_y))
                    
                if not coordinates:
                    print("No CAD_NX found.")
                    self.download_status = 2
                    return False
                else:
                    print(f"{len(coordinates)} item(s) found. Clicking them in turn.")
                    for coord in coordinates:
                        teamcenter_window.set_focus()
                        time.sleep(1)
                        click(coord[0], coord[1])
                        time.sleep(0.1)  # pause between clicks
                        kb.send_keys('{ENTER}')
                        self.total_open_NX += 1
                        if self.stop_flag: 
                            self.update_status.emit(
                                "",
                                "<html><head/><body><p align='center' style='Stopped!</p></body></html>"
                            )
                            return      
                        if self.confirm_window_openned(teamcenter_app, teamcenter_window):
                            if self.app_nx == None:
                                self.app_nx = Application(backend="uia").connect(title_re=".*NX.*", timeout=900)
                                self.window_nx = self.app_nx.window(title_re=".*NX.*")
                                
                                self.export_window = self.window_nx.child_window(title="Export PDF", control_type="Pane")

                            self.window_nx.wait('ready', timeout=999)
                            self.window_nx.set_focus()
                            
                            if self.stop_flag: 
                                self.update_status.emit(
                                    "",
                                    "<html><head/><body><p align='center' style='Stopped!</p></body></html>"
                                )
                                return
                            export_status = export_pdf()
                            
                            if export_status: self.download_status = 1
                            else: self.download_status = 0
                            if self.download_status == 0: 
                                print("Export Error")    
                                continue

            except Exception as e:
                print("Cannot click TreeItem:", e)
                self.download_status = 0
                return 

        if file_type == 'excel' or file_type == 'zip':
            excel_and_zip()
        if file_type == 'nx':
            find_and_open_nx()

        self.waiting_progress(teamcenter_window)

    def main(self, teamcenter_app, teamcenter_window, input_file_1, input_file_2, outputfolder, search_type, coliteam, colrevision, colnamefolder):
        self.total_open_NX = 0
        self.total_turn = 0
        self.first_turn = False
        self.download_status = None  # 0: download error, 1: download success, 2: not found, 3: check again
        self.name_file_log = "download_status"
        col_mapping = {
            "Ref Drawing": "E",
            "Data Note": "D",
            "PDF CAD": "F"
        }
        status_mapping = {
            0: "Download Error",
            1: "Download Success",
            2: "Not Found",
            3: "Check Again"
        }

        def download_type(type):
            i=0
            self.first_turn = True
            self.update_status.emit(
                    f"<html><body><p align='center' style=' font-weight: bold;'>Progress download {type}: {i}/{len(data)} (0%)</p></body></html>",
                    ""
                )
            for idx, (x, y, z, a, b, c) in enumerate(data):
                if self.stop_flag: break
                print(f"--------------------------------------------\n {y} {z}")
                
                self.download_status = None
                if self.total_turn % 50 == 0 and self.total_turn >= 1: self.reset(teamcenter_window)

                i += 1
                self.total_turn += 1
                progress_percentage = (i) / len(data) * 100
                self.update_status.emit(
                    f"<html><body><p align='center' style=' font-weight: bold;'>Progress download {type}: {i}/{len(data)} ({progress_percentage:.1f}%)</p></body></html>",
                    ""
                )
                if y == "NONE" or not z[-1].isdigit():
                    print("Skip None")
                    continue

                if type == "Data Note" and (a == None or a == "Download Error" or a == "Unknown"): self.download_file(teamcenter_app, teamcenter_window, outputfolder,x, y, z, 'excel')
                elif type == "Ref Drawing" and (b == None or b == "Download Error" or b == "Unknown"): self.download_file(teamcenter_app, teamcenter_window, outputfolder,x, y, z, 'zip')
                elif type == "PDF CAD" and (c == None or c == "Download Error" or c == "Unknown"): self.download_file(teamcenter_app, teamcenter_window, outputfolder,x, y, z, 'nx')
                else: 
                    print(f"Skip Download {type}")
                    continue

                col = col_mapping.get(type)
                if col: write_to_cell(os.path.join(outputfolder, f"{self.name_file_log}.xlsx"), idx + 2, col, status_mapping.get(self.download_status, "Unknown"))

        def shorten_list(import_list):
            # Shorten list_temp by merging tuples with identical Shape No and Shape Revision
            merged_data = {}
            for unit, shape_no, shape_rev, stt_datanote, stt_rev, stt_pdf in import_list:
                key = (shape_no, shape_rev, stt_datanote, stt_rev, stt_pdf)
                if key in merged_data:
                    merged_data[key].append(unit)
                else:
                    merged_data[key] = [unit]

            # Build a new list_temp with merged unit names for matching shape_no and shape_rev
            new_list = []
            for key, units in merged_data.items():
                shape_no, shape_rev, stt_datanote, stt_rev, stt_pdf = key
                merged_units = ",".join(sorted(set(units)))
                new_list.append((merged_units, shape_no, shape_rev, stt_datanote, stt_rev, stt_pdf))
            return new_list
        
        def write_into_excel(data_list, output_folder, save_name="output"):
            save_path = os.path.join(output_folder, f"{save_name}.xlsx")
            if os.path.exists(save_path):
                os.remove(save_path)

            wb = Workbook()
            ws = wb.active
            ws.title = save_name

            # Write header row
            headers = ["Unit", "Shape No", "Shape Revision", "Data Note", "Rev Drawing", "NX PDF"]
            ws.append(headers)

            # Write each tuple from data_list into subsequent rows.
            for row_data in data_list:
                ws.append(row_data)

            wb.save(save_path)

        def write_to_cell(file_path, row, col, value):
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb[wb.sheetnames[-1]]
                if isinstance(col, int):
                    col = get_column_letter(col)
                ws[f"{col}{row}"] = value
                wb.save(file_path)
            except Exception as e:
                print(f"Error writing to cell in file {file_path} at row {row} and column {col}: {e}")
                return
        #_________________________________________________________
        # Close Excel window named "download_status.xlsx" if it is open
        for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
            try:
                if proc.info['name'] == 'EXCEL.EXE':
                    cmdline = ' '.join(proc.info.get('cmdline', []))
                    if f'{self.name_file_log}.xlsx' in cmdline:
                        os.kill(proc.info['pid'], signal.SIGTERM)
            except Exception:
                pass

        if not self.ui.file_mode_toggle.isChecked():
            data = self.get_data_from_simple_file(input_file_1, colnamefolder, coliteam, colrevision)
        else:
            data = self.get_data_from_map_file(input_file_1, input_file_2)
            
        data = shorten_list(data)
        write_into_excel(data, outputfolder, self.name_file_log)

        operations = {
            'excel': "Data Note",
            'zip': "Ref Drawing",
            'nx': "PDF CAD"
        }

        self.reset(teamcenter_window)
        for op in search_type:
            label = operations[op]
            self.preparing(teamcenter_window,op)
            download_type(label)

            if self.stop_flag:
                self.update_status.emit(
                    "",
                    "<html><head/><body><p align='center' style='Stopped!</p></body></html>"
                )
            else:
                self.update_status.emit(
                    f"<html><body><p align='center' style=''>Download {label} has finished!</p></body></html>",
                    ""
                )
        self.enable_download_button.emit(True)
        self.update_status.emit(
            "",
            "<html><head/><body><p align='center' style=''>Done</p></body></html>"
        )

        self.end_time = datetime.now().replace(microsecond=0)
        total_time = self.end_time - self.start_time

        print("----------------------------------------------------------")
        print(f"Start at: {self.start_time}")
        print(f"End at: {self.end_time}")
        print(f"Total Running Time: {total_time}")
        
    def main_function(self, input_file_1, input_file_2, outputfolder, coliteam, colrevision, colnamefolder, operation_type):
        if self.stop_flag: return
        app_path = "javaw.exe"
        self.app_nx = None
        self.window_nx = None
        self.export_window = None
        while True:
            if self.stop_flag: return
            try:
                app_teamcenter = Application(backend="uia").connect(path=app_path, timeout= 0.5)
                window_teamcenter = app_teamcenter.window(found_index=0)
                if window_teamcenter.exists():
                    break
            except:
                self.update_status.emit(
                    "",
                    "<html><head/><body><p align='center' style=''>Please Login Teamcenter!</p></body></html>"
                )
                self.show_message.emit("Error", "Please Login Teamcenter!", "error")
                self.enable_download_button.emit(True)
                return

        window_teamcenter.set_focus()
        window_teamcenter.maximize()
        
        self.main(app_teamcenter, window_teamcenter, input_file_1, input_file_2, outputfolder, operation_type, coliteam, colrevision, colnamefolder)

    def download(self):
        self.start_time = datetime.now().replace(microsecond=0)
        
        self.enable_download_button.emit(False)
        input_file_1 = self.ui.input_file_entry_1.text()
        input_file_2 = self.ui.input_file_entry_2.text()
        output_folder = self.ui.output_folder_entry.text()
        coliteam = self.ui.item_column_entry.text()
        colrevision = self.ui.rev_column_entry.text()
        colnamefolder = self.ui.folder_column_entry.text()
        input_var = self.ui.file_mode_toggle.isChecked()
        datanote = self.ui.datanote.isChecked()
        rev_drawing = self.ui.rev_drawing.isChecked()
        nx_pdf = self.ui.nx_pdf.isChecked()

        def is_alpha(value):
            return re.match("^[A-Za-z]+$", value) is not None

        if not input_file_1 and not input_var or not input_file_2 and input_var or not output_folder:
            self.enable_download_button.emit(True)
            self.show_message.emit("Error", "Please select all required files and output folder.", "error")
            return

        if not (is_alpha(coliteam) and is_alpha(colrevision) and is_alpha(colnamefolder)) and not input_var:
            self.enable_download_button.emit(True)
            self.show_message.emit("Error", "Please enter alphabetic characters for the columns.", "error")
            return
        # Check if input files exist
        if not input_var:
            if not os.path.isfile(input_file_1):
                self.enable_download_button.emit(True)
                self.show_message.emit("Error", "Input file does not exist.", "error")
                return
        else:
            if not os.path.isfile(input_file_1) or not os.path.isfile(input_file_2):
                self.enable_download_button.emit(True)
                self.show_message.emit("Error", "One or both MAP/Connector IF files do not exist.", "error")
                return
        # Build operation types based on checkboxes
        op_types = []
        if datanote:
            op_types.append('excel')
        if rev_drawing:
            op_types.append('zip')
        if nx_pdf:
            op_types.append('nx')

        if not op_types:
            self.enable_download_button.emit(True)
            self.show_message.emit("Error", "Please select file type to download.", "error")
            return
        
        self.update_status.emit(
            "<html><head/><body><p align='center'><span style='font-size:10pt;'>Processing...</span></p></body></html>",
            "<html><head/><body><p align='center'><span style='font-size:10pt;'>Running...</span></p></body></html>"
        )
        self.stop_flag = False
        
        def thread_func():
            try:
                self.main_function(input_file_1, input_file_2, output_folder, coliteam, colrevision, colnamefolder, op_types)
                self.update_status.emit(
                    "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt;\">Done</span></p></body></html>",
                    "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt;\">Completed</span></p></body></html>"
                )
                self.show_message.emit("Done", "Download completed!", "info")
            except Exception as e:
                self.update_status.emit(
                    "<html><head/><body><p align=\"center\" style=\"font-size:10pt;\"><span style=\"font-size:10pt;\">Error</span></p></body></html>",
                    "<html><head/><body><p align=\"center\" style=\"font-size:10pt;\"><span style=\"font-size:10pt;\">  </span></p></body></html>"
                )
                self.show_message.emit("Error", str(e), "error")
            finally:
                self.enable_download_button.emit(True)

        threading.Thread(target=thread_func).start()
    
    def stop_task(self):
        self.stop_flag = True
    
        self.update_status.emit(
            "<html><head/><body><p align='center'><span style='font-size:10pt;'>Stopped</span></p></body></html>",
            "<html><head/><body><p align='center'><span style='font-size:10pt;'>Stopped</span></p></body></html>"
        )
        self.enable_download_button.emit(True)

if __name__ == "__main__":
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    if not os.path.exists("log"):
        os.makedirs("log")
    sys.stdout = Logger(os.path.join("log", f"log_{now}.txt"))
    sys.stderr = sys.stdout  # Redirect stderr to log as well
    # Hide the console window if running as a script
    try:
        ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)
    except Exception:
        pass

    app = QApplication(sys.argv)
    window = MainWindow()
    downloader = TeamcenterDownloader(window)
    window.show()
    app.exec_()
