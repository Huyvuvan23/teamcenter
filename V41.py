import os
import re
import time
from pywinauto.application import Application
import sys
import pyperclip
from openpyxl import load_workbook
import xlwings as xw
import openpyxl
from openpyxl.utils import get_column_letter
from pywinauto import keyboard as kb
import keyboard
import signal
import shutil
import win32api, win32con
import pyautogui as pag
import threading
import tkinter as tk
from tkinter import messagebox
from PyQt5.QtWidgets import QApplication, QFileDialog
import psutil
from datetime import datetime, timedelta
import json
from openpyxl import Workbook
import ctypes
import xml.etree.ElementTree as ET

# Redirect stdout to both terminal and file
class Logger:
    def __init__(self, filename):
        self.terminal = sys.stdout
        self.log = open(filename, "w")

    def write(self, message):
        self.terminal.write(message)  # Print to terminal
        self.terminal.flush()         # Force terminal update
        self.log.write(message)      # Write to file
        self.log.flush()             # Force file write (no buffer delay)

    def flush(self):
        pass  # Needed for Python 3 compatibility

class TeamcenterDownloader:
    def __init__(self):
        self.start_time = 0
        self.end_time = 0
        self.data_file = "form_data.json"
        self.data = {}
        self.stop_flag = False
        if os.path.exists(self.data_file):
            with open(self.data_file, 'r') as f:
                self.data = json.load(f)
        self.setup_gui()
        self.load_data()

    def setup_gui(self):
        self.root = tk.Tk()
        self.root.title("Teamcenter File Downloader Tool")
        self.root.configure(bg="#f0f0f0")
        self.root.resizable(False, False)
        self.root.update_idletasks()
        self.root.eval('tk::PlaceWindow . center')
        
        # Create GUI elements
        # self.create_warning_label()
        self.create_io_frame()
        self.settings()
        self.choose_file_type_frame()
        self.create_progress_frame()
        self.create_button_frame()
        
        self.update_visibility()
        # Set window close handler
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def create_warning_label(self):
        self.warning_label = tk.Label(self.root, 
            text="Do not operate while the tool is running.\nLog in to Teamcenter before running.",
            fg="red", bg="#f0f0f0", wraplength=500, justify="center")
        self.warning_label.pack(pady=10)

    def create_io_frame(self):
        self.io_frame = tk.Frame(self.root, bg="#f0f0f0")
        self.io_frame.pack(pady=5, padx=5, fill="x")
        self.io_frame.grid_columnconfigure(0, weight=1)
        self.io_frame.grid_columnconfigure(1, weight=1)
        self.io_frame.grid_columnconfigure(2, weight=1)

        # Store input_file_label_1 as instance variable
        self.input_file_label_1 = tk.Label(self.io_frame, text="Input File:", bg="#f0f0f0", width=15, anchor="e")
        self.input_file_label_1.grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.input_file_entry_1 = tk.Entry(self.io_frame, width=70)
        self.input_file_entry_1.grid(row=0, column=1, columnspan=2, padx=5, pady=5, sticky="nsew")
        self.input_file_button_1 = tk.Button(self.io_frame, text="Browse", command=lambda: self.select_input_file(self.input_file_entry_1, "MAP" if self.input_file_label_1.cget("text").lower().startswith("map") else ""))
        self.input_file_button_1.grid(row=0, column=3, padx=5, pady=5, sticky="nsew")

        # Store input_file_label_2 as instance variable
        self.input_file_label_2 = tk.Label(self.io_frame, text="Connector IF File:", bg="#f0f0f0", width=15, anchor="e")
        self.input_file_label_2.grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.input_file_entry_2 = tk.Entry(self.io_frame, width=70)
        self.input_file_entry_2.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="nsew")
        self.input_file_button_2 = tk.Button(self.io_frame, text="Browse", command=lambda: self.select_input_file(self.input_file_entry_2, "Connector IF"))
        self.input_file_button_2.grid(row=1, column=3, padx=5, pady=5, sticky="nsew")

        output_folder_label = tk.Label(self.io_frame, text="Output Folder:", bg="#f0f0f0", width=15, anchor="e")
        output_folder_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.output_folder_entry = tk.Entry(self.io_frame, width=70)
        self.output_folder_entry.grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky="nsew")
        output_folder_button = tk.Button(self.io_frame, text="Browse", command=lambda: self.select_output_folder(self.output_folder_entry))
        output_folder_button.grid(row=2, column=3, padx=5, pady=5, sticky="nsew")

    def settings(self):
        self.settings_frame = tk.Frame(self.root, bg="#f0f0f0")
        self.settings_frame.pack(pady=5, padx=5, fill="x")

        self.column_frame = tk.LabelFrame(self.settings_frame, text="Settings", padx=10, pady=10, bg="#f0f0f0")
        self.column_frame.pack(pady=10, fill="both", expand=True)

        # Configure grid columns for proper resizing
        for col in range(3):
            self.column_frame.grid_columnconfigure(col, weight=1)

        # Column for Folders
        self.colnamefolder_label = tk.Label(self.column_frame, text="Column for Folders:", bg="#f0f0f0")
        self.colnamefolder_label.grid(row=0, column=1, padx=5, pady=5, sticky="e")
        self.colnamefolder_entry = tk.Entry(self.column_frame, width=5)
        self.colnamefolder_entry.grid(row=0, column=2, padx=5, pady=5, sticky="w")

        # Column for Item ID
        self.coliteam_label = tk.Label(self.column_frame, text="Column for Item ID:", bg="#f0f0f0")
        self.coliteam_label.grid(row=1, column=1, padx=5, pady=5, sticky="e")
        self.coliteam_entry = tk.Entry(self.column_frame, width=5)
        self.coliteam_entry.grid(row=1, column=2, padx=5, pady=5, sticky="w")

        # Column for Revision
        self.colrevision_label = tk.Label(self.column_frame, text="Column for Revision:", bg="#f0f0f0")
        self.colrevision_label.grid(row=2, column=1, padx=5, pady=5, sticky="e")
        self.colrevision_entry = tk.Entry(self.column_frame, width=5)
        self.colrevision_entry.grid(row=2, column=2, padx=5, pady=5, sticky="w")

        # Visibility Selection

        self.input_file_var = tk.IntVar(value=self.data.get("var", 1))
        map_file_radio = tk.Radiobutton(self.column_frame, text="Using MAP File", variable=self.input_file_var, value=0, command=self.update_visibility, bg="#f0f0f0")
        simple_file_radio = tk.Radiobutton(self.column_frame, text="Using Simple Input File", variable=self.input_file_var, value=1, command=self.update_visibility, bg="#f0f0f0")
        map_file_radio.grid(row=2, column=0, padx=50, pady=5, sticky="w")
        simple_file_radio.grid(row=1, column=0, padx=50, pady=5, sticky="w")

    def update_visibility(self):
        if self.input_file_var.get() == 1:
            self.input_file_label_2.grid_remove()
            self.input_file_entry_2.grid_remove()
            self.input_file_button_2.grid_remove()

            self.colnamefolder_label.grid()
            self.colnamefolder_entry.grid()
            self.coliteam_label.grid()
            self.coliteam_entry.grid()
            self.colrevision_label.grid()
            self.colrevision_entry.grid()

            self.input_file_label_1.config(text="Input File:")

        else:
            self.input_file_label_2.grid()
            self.input_file_entry_2.grid()
            self.input_file_button_2.grid()

            self.colnamefolder_label.grid_remove()
            self.colnamefolder_entry.grid_remove()
            self.coliteam_label.grid_remove()
            self.coliteam_entry.grid_remove()
            self.colrevision_label.grid_remove()
            self.colrevision_entry.grid_remove()

            self.input_file_label_1.config(text="MAP File:")

    def choose_file_type_frame(self):
        self.choose_file_type = tk.LabelFrame(self.settings_frame,  text="", padx=10, pady=10, bg="#f0f0f0")
        self.choose_file_type.pack(pady=5)

        # Create IntVar variables for each checkbox
        self.data_note_var = tk.IntVar(value=0)
        self.ref_drawing_var = tk.IntVar(value=0)
        self.pdf_cad_var = tk.IntVar(value=0)

        # Create checkboxes and pack them into the frame
        data_note_cb = tk.Checkbutton(self.choose_file_type, text="Data Note", variable=self.data_note_var, bg="#f0f0f0")
        ref_drawing_cb = tk.Checkbutton(self.choose_file_type, text="Ref Drawing", variable=self.ref_drawing_var, bg="#f0f0f0")
        pdf_cad_cb = tk.Checkbutton(self.choose_file_type, text="PDF CAD", variable=self.pdf_cad_var, bg="#f0f0f0")

        data_note_cb.pack(side=tk.LEFT, padx=5)
        ref_drawing_cb.pack(side=tk.LEFT, padx=5)
        pdf_cad_cb.pack(side=tk.LEFT, padx=5)


    def create_progress_frame(self):
        self.progress_frame = tk.Frame(self.root, bg="#f0f0f0")
        self.progress_frame.pack(pady=5)

        self.progress_label = tk.Label(self.progress_frame, text="", fg="blue", bg="#f0f0f0")
        self.progress_label.pack()

        self.done_label = tk.Label(self.progress_frame, text="", fg="green", bg="#f0f0f0")
        self.done_label.pack()

    def create_button_frame(self):
        self.button_frame = tk.Frame(self.root, bg="#f0f0f0")
        self.button_frame.pack(pady=10)

        self.download_button = tk.Button(self.button_frame, text="Download", command=self.download, bg="#0B3040", fg="white", width=15, font=("Arial", 10))
        self.download_button.pack(side=tk.LEFT, padx=10)

        self.stop_button = tk.Button(self.button_frame, text="Stop", command=self.stop_task, bg="#f44336", fg="white", width=15, font=("Arial", 10))
        self.stop_button.pack(side=tk.LEFT, padx=10)

    def save_data(self):
        data = {
            "input_file_1": self.input_file_entry_1.get(),
            "input_file_2": self.input_file_entry_2.get(),
            "output_folder": self.output_folder_entry.get(),
            "coliteam": self.coliteam_entry.get(),
            "colrevision": self.colrevision_entry.get(),
            "colnamefolder": self.colnamefolder_entry.get(),
            "var": self.input_file_var.get(),
        }
        with open(self.data_file, 'w') as f:
            json.dump(data, f)

    def load_data(self):
        self.input_file_entry_1.insert(0, self.data.get("input_file_1", ""))
        self.input_file_entry_2.insert(0, self.data.get("input_file_2", ""))
        self.output_folder_entry.insert(0, self.data.get("output_folder", ""))
        self.coliteam_entry.insert(0, self.data.get("coliteam", ""))
        self.colrevision_entry.insert(0, self.data.get("colrevision", ""))
        self.colnamefolder_entry.insert(0, self.data.get("colnamefolder", ""))

    def select_input_file(self, entry_widget, text):
        app = QApplication(sys.argv)
        file_path, _ = QFileDialog.getOpenFileName(None, f"Select File {text}", None, "Excel Files (*.xls*);;All Files (*)")
        app.quit()

        if file_path:
            file_path = file_path.replace("/", "\\")
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, file_path)

    def select_output_folder(self, entry_widget):
        app = QApplication(sys.argv)
        folder_path = QFileDialog.getExistingDirectory(None, "Select Folder")
        app.quit()

        if folder_path:
            folder_path = folder_path.replace("/", "\\")
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, folder_path)

    def waiting_progress(self, teamcenter_window):
        time_start = time.time()
        while True:
            if self.stop_flag: return
            if time.time() - time_start >= 5*60:
                self.done_label.config(text="Error, Please try again", fg="red")
                print("Out of waiting time")
                sys.exit()
            if teamcenter_window.child_window(control_type="Text", title="No operations to display at this time.").exists():
                return False

    def reset(self, teamcenter_window):
        if self.stop_flag: return 
        teamcenter_window.set_focus()
        kb.send_keys("%WOM")
        self.waiting_progress(teamcenter_window)
        kb.send_keys("%WRY")
        self.waiting_progress(teamcenter_window)
        try:
            teamcenter_window.child_window(title="Close All", control_type="MenuItem").select()
        except:
            pass

    def preparing(self, teamcenter_window, type):
        if self.stop_flag: return
        teamcenter_window.child_window(control_type="SplitButton", title="Select a Search").invoke()
        self.waiting_progress(teamcenter_window)

        if type == 'excel':
            kb.send_keys("I{ENTER}")
        if type == 'zip':
            kb.send_keys("R{ENTER}")
        if type == 'nx':    
            kb.send_keys("SSS{ENTER}")

        self.waiting_progress(teamcenter_window)
        teamcenter_window.child_window(control_type="Button", title="Clear all search fields").invoke()

    def get_data_from_simple_file(self, file_link, colnamefolder, coliteam, colrevision):
        data = []
        try:
            wb = load_workbook(filename=file_link)
            sheet = wb.active
            maxrow = sheet.max_row
                
            for i in range(2, maxrow + 1):  # Start from row 2, assuming row 1 is header
                item = sheet[f'{coliteam}{i}'].value
                revision = sheet[f'{colrevision}{i}'].value
                folder_name = sheet[f'{colnamefolder}{i}'].value
                header_values = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                if header_values == ["Unit", "Shape No", "Shape Revision", "Data Note", "Rev Drawing", "NX PDF"]:
                    col_idx = openpyxl.utils.column_index_from_string(colrevision)
                    stt_datanote = sheet.cell(row=i, column=col_idx + 1).value
                    stt_rev = sheet.cell(row=i, column=col_idx + 2).value
                    stt_pdf = sheet.cell(row=i, column=col_idx + 3).value
                else:
                    stt_datanote = ''
                    stt_rev = ''
                    stt_pdf = ''
                # Only add if all fields are non-empty strings
                if all(isinstance(val, str) and val.strip() for val in [item, revision, folder_name]):
                    tup = (folder_name.strip(), item.strip(), revision.strip(), stt_datanote, stt_rev, stt_pdf)
                    if tup not in data:
                        data.append(tup)
        except Exception as e:
            print(f"Error reading input file: {e}")
        return data

    def get_data_from_map_file(self, link_MAP_file, link_connector_infor_file):
        def set_up_ws(ws):
            try:
                ws.api.Rows.Hidden = False
                ws.api.Columns.Hidden = False
                ws.api.AutoFilterMode = False
            except Exception:
                pass

        def get_last_row(ws, col):
            return ws.range(col + str(ws.cells.last_cell.row)).end('up').row

        def find_cell(ws, text_to_find):
            unit_name_cell_ws_if = None
            try:
                found = ws.api.Cells.Find(
                    What = text_to_find,
                    LookIn=-4163,  # xlValues
                    LookAt=1,      # xlWhole
                    SearchOrder=1, # xlByRows
                    SearchDirection=1, # xlNext
                    MatchCase=False
                )
                if found is not None:
                    unit_name_cell_ws_if = ws.range((found.Row, found.Column))
            except Exception:
                pass
            return unit_name_cell_ws_if
        
        app = xw.App(visible=False)
        wb_if = None
        wb_map = None

        try:
            wb_if = app.books.open(link_connector_infor_file)
            ws_if = wb_if.sheets("コネクタIFの作成管理")
            set_up_ws(ws_if)

            unit_name_cell_ws_if = find_cell(ws_if, "unit name")
            
            # Get the last row and all unit names from the "unit name" column in the IF sheet
            if unit_name_cell_ws_if is None:
                raise ValueError("'unit name' cell not found in コネクタIFの作成管理 sheet.")
            unit_name_cell_col_if = get_column_letter(unit_name_cell_ws_if.column)
            unit_name_cell_row_if = unit_name_cell_ws_if.row
            last_row_if = get_last_row(ws_if, unit_name_cell_col_if)
            unitnames_if = ws_if.range(f"{unit_name_cell_col_if}{unit_name_cell_row_if + 3}:{unit_name_cell_col_if}{last_row_if}").value
            unitnames_if = sorted(set(x for x in unitnames_if if x is not None))

            wb_map = app.books.open(link_MAP_file)
            ws_map = wb_map.sheets("MAP")
            ws_dwg = wb_map.sheets["DWG"]
            set_up_ws(ws_map)
            set_up_ws(ws_dwg)
            
            unit_name_cell_map = find_cell(ws_map, "unit name")
            part_number_cell_map = find_cell(ws_map, "Part#(10)")

            # Use the column letter for "unit name" dynamically
            unit_name_cell_col_map = get_column_letter(unit_name_cell_map.column)
            unit_name_cell_row_map = unit_name_cell_map.row
            last_row_map = get_last_row(ws_map, unit_name_cell_col_map)
            unitname_map = ws_map.range(f"{unit_name_cell_col_map}{unit_name_cell_row_map + 1}:{unit_name_cell_col_map}{last_row_map}").value

            part_number_cell_col_map = get_column_letter(part_number_cell_map.column)
            part_number_map = ws_map.range(f"{part_number_cell_col_map}{unit_name_cell_row_map + 1}:{part_number_cell_col_map}{last_row_map}").value

            no_cell_dwg = find_cell(ws_dwg, "no")
            no_cell_col_dwg = get_column_letter(no_cell_dwg.column)
            no_cell_row_dwg = no_cell_dwg.row
            last_row_dwg = get_last_row(ws_dwg, no_cell_col_dwg)
            
            # Get the columns by checking the header row
            headers = ws_dwg.range(f"A{no_cell_row_dwg}:XFD{no_cell_row_dwg}").value
            try:
                part_number_column = headers.index("PART #") + 1
                shape_no_column = headers.index("Drawing/Shape No") + 1
                shape_revision_column = headers.index("Drawing/Shape Revision") + 1
            except ValueError as e:
                raise ValueError("One or more necessary column headers not found.") from e
            
            data_start = no_cell_row_dwg + 1
            list_part_number_dwg = ws_dwg.range(ws_dwg.cells(data_start, part_number_column),
                                                ws_dwg.cells(last_row_dwg, part_number_column)).value
            list_shape_no_dwg = ws_dwg.range(ws_dwg.cells(data_start, shape_no_column),
                                            ws_dwg.cells(last_row_dwg, shape_no_column)).value
            list_shape_revision_dwg = ws_dwg.range(ws_dwg.cells(data_start, shape_revision_column),
                                                ws_dwg.cells(last_row_dwg, shape_revision_column)).value
            
            list_shape_merged_dwg = [
                f"{no_val},{rev_val}" if no_val is not None and rev_val is not None else ""
                for no_val, rev_val in zip(list_shape_no_dwg, list_shape_revision_dwg)
            ]
            
            # Group part numbers based on unit names from the "IF" sheet.
            part_numbers_by_unit = []
            for unit in unitnames_if:
                indices = [i for i, nm in enumerate(unitname_map) if nm == unit]
                values = sorted(set(part_number_map[i] for i in indices if part_number_map[i] is not None))
                part_numbers_by_unit.append(values)
            
            # Get the shapes for each unit
            shapes_by_unit = []
            for part_numbers in part_numbers_by_unit:
                unit_shapes = []
                for part in part_numbers:
                    indices = [i for i, pn in enumerate(list_part_number_dwg) if pn == part]
                    unit_shapes.extend(list_shape_merged_dwg[i] for i in indices)
                shapes_by_unit.append(sorted(set(unit_shapes)))
            
            all_data = []
            for idx, unit_shapes in enumerate(shapes_by_unit):
                is_ok = False
                for item in unit_shapes:
                    if item:
                        is_ok = True
                        shape_parts = item.split(',')
                        if len(shape_parts) == 2:
                            all_data.append((unitnames_if[idx], shape_parts[0], shape_parts[1], '', '', ''))
                if not is_ok:
                    all_data.append((unitnames_if[idx], "NONE", "NONE", '', '', ''))
            
            return all_data
        finally:
            if wb_if:
                wb_if.close()
            if wb_map:
                wb_map.close()
            app.quit()

    def get_latest_excel_zip_file(self):
        folder_path = rf'C:\Temp'
        folders = [f for f in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, f))]
        if not folders:
            return None
        
        full_folder_paths = [os.path.join(folder_path, f) for f in folders]
        latest_folder = max(full_folder_paths, key=os.path.getmtime)
        
        excel_files = [f for f in os.listdir(latest_folder) if (f.endswith('.xlsm') or f.endswith('.xls') or f.endswith('.zip') or f.endswith('.xlsx'))  and not f.startswith('~$')]
        if not excel_files:
            return None
        
        full_file_paths = [os.path.join(latest_folder, f) for f in excel_files]
        latest_file = max(full_file_paths, key=os.path.getmtime)
        
        return latest_file

    def copy_latest_excel_zip_file(self, new_folder_moi, tenmoi):
        latest_file = self.get_latest_excel_zip_file()
        if latest_file:
            destination_folder = new_folder_moi
            destination_path = os.path.join(destination_folder, os.path.basename(latest_file))
            
            base, extension = os.path.splitext(destination_path)
            destination_path = f"{base}_{tenmoi}{extension}"
            counter = 1
            original_base = base
            while os.path.exists(destination_path):
                destination_path = f"{original_base}_{tenmoi}_{counter}{extension}"
                counter += 1
            
            shutil.copy2(latest_file, destination_path)
            print(f"Latest file '{os.path.basename(latest_file)}' copied to '{destination_folder}'.")
        else:
            print("No file found.")

    def kill_new_excel_processes(self):
        threshold_time = datetime.now() - timedelta(minutes=2)
        for proc in psutil.process_iter(['pid', 'name', 'create_time']):
            if proc.info['name'] == 'EXCEL.EXE':
                process_create_time = datetime.fromtimestamp(proc.info['create_time'])
                if process_create_time > threshold_time:
                    os.kill(proc.info['pid'], signal.SIGTERM)

    def kill_new_7zip_processes(self):
        threshold_time = datetime.now() - timedelta(minutes=2)
        for proc in psutil.process_iter(['pid', 'name', 'create_time']):
            if proc.info['name'] =='7zFM.exe':
                process_create_time = datetime.fromtimestamp(proc.info['create_time'])
                if process_create_time > threshold_time:
                    os.kill(proc.info['pid'], signal.SIGTERM)

    def confirm_window_openned(self, teamcenter_app, teamcenter_window):
        star_time = time.time()
        confirm_window = teamcenter_window.child_window(class_name="SunAwtDialog")
        while True:
            if self.stop_flag: return
            windows = teamcenter_app.windows()
            for window in windows:
                try:
                    if window != teamcenter_window:
                        window.set_focus()
                        time.sleep(1)
                        kb.send_keys("{ENTER}")
                    else:
                        if confirm_window . exists():
                            confirm_window . child_window(title="Close", control_type="Button").invoke()
                            return False
                except:
                    pass
            if time.time() - star_time >= 300:
                print("Out of waiting time")
                sys.exit()
            
            else:
                if teamcenter_window.child_window(control_type="Text", title="No operations to display at this time.").exists():
                    return True
                else:
                    continue

    def create_folder(self, outputfolder, folder_name):
        # Replace all characters not allowed in folder names with "_"
        sanitized_folder_name = re.sub(r'[<>:"/\\|?*]', '_', str(folder_name))
        folder_path = os.path.join(outputfolder, sanitized_folder_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        return folder_path

    def set_search_fields(self, search_window, item_id, revision, file_type):
        """
        Set the search fields in the Teamcenter search window based on file_type.
        Args:
            search_window: pywinauto window object for the search tab.
            item_id: The item ID or shape number to search for.
            revision: The revision or shape change number.
            file_type: One of 'excel', 'zip', or 'nx'.
        """
        if file_type == 'excel':
            # For Data Note (Excel), set Item ID and Revision fields
            search_window.child_window(control_type="Edit", title="Item ID:").set_text(item_id)
            search_window.child_window(control_type="Edit", title="Revision:").set_text(revision)
        elif file_type == 'zip':
            # For Ref Drawing (ZIP), set Shape Number and Shape Change Number fields
            search_window.child_window(control_type="Edit", title="Shape Number:").set_text(item_id)
            search_window.child_window(control_type="Edit", title="Shape Change Number:").set_text(revision)
        elif file_type == 'nx':
            # For NX PDF, handle ShapeNumber and ShapeName fields
            shapenumber = search_window.child_window(control_type="Edit", title="ShapeNumber:")
            shapename = search_window.child_window(control_type="Edit", title="ShapeName:")
            btn_more = search_window.child_window(title="More...>>>", control_type="Button")
            i = 0
            # If item_id ends with '*', use ShapeNumber, else use ShapeName
            if item_id.endswith("*"):
                # Ensure ShapeNumber field is visible
                while not shapenumber.exists() and i < 10:
                    btn_more.click()
                    time.sleep(0.5)
                    i += 1
                    shapename.set_text("")
                if not shapenumber.exists() : self.preparing(search_window.parent(), "nx")
                shapenumber.set_text(item_id)
            else:
                # Ensure ShapeName field is visible
                while not shapename.exists() and i < 10:
                    btn_more.click()
                    time.sleep(0.5)
                    i += 1
                    shapenumber.set_text("")
                if not shapename.exists() : self.preparing(search_window.parent(), "nx")
                shapename.set_text(item_id)
            # Set ShapeChangeNumber field
            search_window.child_window(control_type="Edit", title="ShapeChangeNumber:").set_text(revision)

    def download_file(self, teamcenter_app, teamcenter_window, outputfolder, folder_name, item_id, revision, file_type):
        
        if file_type == 'excel':
            item_id_moi = item_id + "-note"
        elif file_type == 'nx':
            item_id_moi = item_id + "-shape"
        else:
            item_id_moi = item_id
        revision_moi = "0" + revision[-2:]
        
        search_window = teamcenter_window.child_window(title="Search", control_type="Tab")
        self.set_search_fields(search_window, item_id_moi, revision_moi, file_type)
        kb.send_keys("{ENTER}")
        self.waiting_progress(teamcenter_window)

        kiem_tra_item_window = teamcenter_window.child_window(title="Search Results", control_type="Tab")
        
        if file_type == 'excel':
            if kiem_tra_item_window.child_window(control_type="Pane", title="Item Revision...  - No objects found").exists():
                print("No object found for Excel")
                self.download_status = 3
                return False
        elif file_type == 'zip':
            if kiem_tra_item_window.child_window(control_type="Pane", title="Ref Drawing  - No objects found").exists():
                print("No ZIP file found")
                self.download_status = 2
                return False
        
        elif file_type == 'nx':
            status = kiem_tra_item_window.child_window(control_type="Pane", title="Shape  - No objects found")
            if status.exists():
                self.set_search_fields(search_window, item_id + "*", revision_moi, file_type)
                kb.send_keys("{ENTER}")
                self.waiting_progress(teamcenter_window)
                if status.exists():
                    print("No Shape found")
                    self.download_status = 3
                    return False

        pane_window = kiem_tra_item_window.child_window(control_type="Pane", title="Search Results")

        def excel_and_zip():
            try:
                pane_window.child_window(control_type="TreeItem", found_index=1).select()
            except:
                print("Cannot click TreeItem 1")
                self.download_status = 0
                return False

            self.waiting_progress(teamcenter_window)
            pane_window.child_window(control_type="Button", title="", found_index=4).click()

            for index in range(3, 11):
                file_ez = kiem_tra_item_window.child_window(control_type="TreeItem", found_index=index)
                if file_ez.exists():
                    try:
                        file_ez.type_keys("{ENTER}")
                    except Exception as e: 
                        print(f"Have Error: {e}")
                        self.download_status = 0
                        return
                    
                    if self.confirm_window_openned(teamcenter_app, teamcenter_window):
                        time.sleep(1)
                        for each_folder in folder_name.split(","):
                            folder_path = self.create_folder(outputfolder, each_folder)
                            self.copy_latest_excel_zip_file(folder_path, revision_moi)

                        if file_type == 'excel':
                            self.kill_new_excel_processes()
                        else:
                            self.kill_new_7zip_processes()
                        time.sleep(1)
                        self.download_status = 1
                else:   
                    print(f"No file found at position {index}")
                    break
        
        def export_pdf():
            def name_pdf_cad(file_xml):
                # Parse the XML file
                tree = ET.parse(file_xml)
                root = tree.getroot()

                # Define the XML namespace (found in the root tag)
                ns = {'plm': 'http://www.plmxml.org/Schemas/PLMXMLSchema'}

                # Find the Representation element and get its 'name' attribute
                representation = root.find('.//plm:Representation', ns)
                if representation is not None:
                    name = representation.get('name')
                    return name
                else:
                    return 0
                
            if self.stop_flag: 
                self.done_label.config(text="Stopped!", fg="red")
                return
            name_nx_cad = name_pdf_cad(r"C:\Temp\NX_Nav_.plmxml")
            filename = f"{name_nx_cad},{revision_moi}.pdf"

            list_name_folder = folder_name.split(",")
            for idx, each_folder in enumerate(list_name_folder):
                try:
                    
                    folder_path = self.create_folder(outputfolder, each_folder)
                    file_path = os.path.join(folder_path, filename)
                    counter = 1
                    while os.path.exists(file_path):
                        filename = f"({counter}) {name_nx_cad},{revision_moi}.pdf"
                        file_path = os.path.join(folder_path, filename)
                        counter += 1
                    if self.stop_flag: 
                        self.done_label.config(text="Stopped!", fg="red")
                        return
                    
                    if idx != 0:
                        shutil.copy2(os.path.join(self.create_folder(outputfolder, list_name_folder[idx-1]), filename), file_path)
                        print(f"File '{os.path.basename(filename)}' is saved to '{each_folder}'.")
                        continue

                    pyperclip.copy(file_path)
                    self.window_nx.set_focus()
                    time.sleep(1)
                    kb.send_keys('%f')
                    for _ in range(15):
                        kb.send_keys('{TAB}')
                        time.sleep(0.02)
                    kb.send_keys('{ENTER}')
                    time.sleep(1)
                    keyboard.send('d')
                    
                    if self.stop_flag: 
                        self.done_label.config(text="Stopped!", fg="red")
                        return
                    self.export_window.wait('ready', timeout=5)
                    if self.first_turn:
                        self.export_window.child_window(control_type="ComboBox",found_index=0).wrapper_object().select("File Browser")
                        kb.send_keys('{TAB}')
                    kb.send_keys('^v')
                    # export_window.child_window(control_type="Edit", found_index=0).set_edit_text(file_path)
                    # Hold Shift, press Tab 4 times, press Down 10 times, then release Shift
                    kb.send_keys('{VK_SHIFT down}')  # Hold Shift down
                    for _ in range(3):
                        kb.send_keys('{TAB}')
                        time.sleep(0.02)
                    kb.send_keys('{VK_SHIFT up}')
                    for _ in range(10):
                        kb.send_keys('{VK_UP}')
                        time.sleep(0.02)
                    kb.send_keys('{VK_DOWN}')
                    kb.send_keys('{VK_SHIFT down}')
                    for _ in range(10):
                        kb.send_keys('{VK_DOWN}')
                        time.sleep(0.02)
                    kb.send_keys('{VK_SHIFT up}')  # Release Shift

                    if self.first_turn:
                        self.export_window.child_window(control_type="ComboBox",found_index=1).wrapper_object().select("Black on White")

                    self.export_window.child_window(title="OK", control_type="Button").click()
                    if self.export_window.exists():
                        kb.send_keys('{ENTER}')
                        kb.send_keys('{VK_ESCAPE}')
                        return False
                        
                    self.window_nx.wait('ready', timeout=999)
                    if self.stop_flag: 
                        self.done_label.config(text="Stopped!", fg="red")
                        return
                    self.first_turn = False
                    print(f"File '{os.path.basename(filename)}' is saved to '{each_folder}'.")
                    
                except Exception as e:
                    print(f"Have error with NX: {e}")
                    self.app_nx.kill()
                    self.app_nx = None
                    return False

            self.window_nx.wait('ready', timeout=999)

            time.sleep(1)
            if self.total_open_NX % 5 == 0:
                kb.send_keys('%f')
                for _ in range(5):
                    kb.send_keys('{TAB}')
                    time.sleep(0.02)
                time.sleep(0.5)
                kb.send_keys('{ENTER}')
                kb.send_keys('{TAB}')
                time.sleep(0.5)
                kb.send_keys('{ENTER}')
                time.sleep(1)
                kb.send_keys('N')
            return True

        def find_and_open_nx():
            teamcenter_window.set_focus()

            def click(x, y):
                """Simulates a single mouse click at the specified (x, y) coordinates."""
                win32api.SetCursorPos((x, y))
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0)
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0)

            try:
                # Select and click up to 3 TreeItems in sequence
                for idx in range(1, 4):
                    try:
                        pane_window.child_window(control_type="TreeItem", found_index=idx).wrapper_object().select()
                        pane_window.child_window(control_type="Button", title="", found_index=4).click()
                        self.waiting_progress(teamcenter_window)
                        if self.stop_flag:
                            self.done_label.config(text="Stopped!", fg="red")
                            return
                    except Exception as e:
                        print(f"Error processing TreeItem {idx}: {e}")
                        break
                # Find all occurrences of "CAD_NX.png" on screen with error handling
                found_images = []
                num = 1
                while True:
                    image_file = f"images/CAD_NX_{num}.png"
                    if not os.path.exists(image_file):
                        break
                    try:
                        images_found = list(pag.locateAllOnScreen(image_file, confidence=0.9))
                        found_images.extend(images_found)
                    except Exception:
                        num += 1
                        continue
                    num += 1

                coordinates = []
                for region in found_images:
                    center_x = region.left + (region.width // 2)
                    center_y = region.top + (region.height // 2)
                    coordinates.append((center_x, center_y))
                    
                if not coordinates:
                    print("No CAD_NX found.")
                    self.download_status = 2
                    return False
                else:
                    print(f"{len(coordinates)} item(s) found. Clicking them in turn.")
                    for coord in coordinates:
                        teamcenter_window.set_focus()
                        time.sleep(1)
                        click(coord[0], coord[1])
                        time.sleep(0.1)  # pause between clicks
                        kb.send_keys('{ENTER}')
                        self.total_open_NX += 1
                        if self.stop_flag: 
                            self.done_label.config(text="Stopped!", fg="red")
                            return      
                        if self.confirm_window_openned(teamcenter_app, teamcenter_window):
                            if self.app_nx == None:
                                self.app_nx = Application(backend="uia").connect(title_re=".*NX.*", timeout=900)
                                self.window_nx = self.app_nx.window(title_re=".*NX.*")
                                self.export_window = self.window_nx.child_window(title="Export PDF", control_type="Pane")
                            self.window_nx.wait('ready', timeout=999)
                            self.window_nx.set_focus()
                            if not self.window_nx.is_maximized():
                                self.window_nx.maximize()
                            kb.send_keys('^+d')
                            self.window_nx.wait('ready', timeout=999)
                            if self.stop_flag: 
                                self.done_label.config(text="Stopped!", fg="red")
                                return
                            export_status = export_pdf()
                            if self.download_status == 0: 
                                print("Export Error")    
                                continue
                            if export_status: self.download_status = 1
                            else: self.download_status = 0

            except Exception as e:
                print("Cannot click TreeItem:", e)
                self.download_status = 0
                return 

        if file_type == 'excel' or file_type == 'zip':
            excel_and_zip()
        if file_type == 'nx':
            find_and_open_nx()

        self.waiting_progress(teamcenter_window)

    def main(self, teamcenter_app, teamcenter_window, input_file_1, input_file_2, outputfolder, search_type, coliteam, colrevision, colnamefolder):
        self.total_open_NX = 0
        self.total_turn = 0
        self.first_turn = False
        self.download_status = None  # 0: download error, 1: download success, 2: not found, 3: check again
        self.name_file_log = "download_status"
        col_mapping = {
            "Ref Drawing": "E",
            "Data Note": "D",
            "PDF CAD": "F"
        }
        status_mapping = {
            0: "Download Error",
            1: "Download Success",
            2: "Not Found",
            3: "Check Again"
        }

        def download_type(type):
            i=0
            self.first_turn = True
            self.progress_label.config(text=f"Progress download {type}: {i}/{len(data)} (0%)")
            for idx, (x, y, z, a, b, c) in enumerate(data):
                if self.stop_flag: break
                print(f"--------------------------------------------\n {y} {z}")
                
                self.download_status = None
                if self.total_turn % 50 == 0 and self.total_turn >= 1: self.reset(teamcenter_window)

                i += 1
                self.total_turn += 1
                progress_percentage = (i) / len(data) * 100
                self.progress_label.config(text=f"Progress download {type}: {i}/{len(data)} ({progress_percentage:.1f}%)")

                if y == "NONE" or not z[-1].isdigit():
                    print("Skip None")
                    continue

                if type == "Data Note" and (a == None or a == "Download Error"): self.download_file(teamcenter_app, teamcenter_window, outputfolder,x, y, z, 'excel')
                elif type == "Ref Drawing" and (b == None or b == "Download Error"): self.download_file(teamcenter_app, teamcenter_window, outputfolder,x, y, z, 'zip')
                elif type == "PDF CAD" and (c == None or c == "Download Error"): self.download_file(teamcenter_app, teamcenter_window, outputfolder,x, y, z, 'nx')
                else: 
                    print(f"Skip Download {type}")
                    continue

                col = col_mapping.get(type)
                if col: write_to_cell(os.path.join(outputfolder, f"{self.name_file_log}.xlsx"), idx + 2, col, status_mapping.get(self.download_status, "Unknown"))

        def shorten_list(import_list):
            # Shorten list_temp by merging tuples with identical Shape No and Shape Revision
            merged_data = {}
            for unit, shape_no, shape_rev, stt_datanote, stt_rev, stt_pdf in import_list:
                key = (shape_no, shape_rev, stt_datanote, stt_rev, stt_pdf)
                if key in merged_data:
                    merged_data[key].append(unit)
                else:
                    merged_data[key] = [unit]

            # Build a new list_temp with merged unit names for matching shape_no and shape_rev
            new_list = []
            for key, units in merged_data.items():
                shape_no, shape_rev, stt_datanote, stt_rev, stt_pdf = key
                merged_units = ",".join(sorted(set(units)))
                new_list.append((merged_units, shape_no, shape_rev, stt_datanote, stt_rev, stt_pdf))
            return new_list
        
        def write_into_excel(data_list, output_folder, save_name="output"):
            save_path = os.path.join(output_folder, f"{save_name}.xlsx")
            if os.path.exists(save_path):
                os.remove(save_path)

            wb = Workbook()
            ws = wb.active
            ws.title = save_name

            # Write header row
            headers = ["Unit", "Shape No", "Shape Revision", "Data Note", "Rev Drawing", "NX PDF"]
            ws.append(headers)

            # Write each tuple from data_list into subsequent rows.
            for row_data in data_list:
                ws.append(row_data)

            wb.save(save_path)

        def write_to_cell(file_path, row, col, value):
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb[wb.sheetnames[-1]]
                if isinstance(col, int):
                    col = get_column_letter(col)
                ws[f"{col}{row}"] = value
                wb.save(file_path)
            except Exception as e:
                print(f"Error writing to cell in file {file_path} at row {row} and column {col}: {e}")
                return
        #_________________________________________________________

        self.reset(teamcenter_window)
        if self.input_file_var.get() == 1:
            data = self.get_data_from_simple_file(input_file_1, colnamefolder, coliteam, colrevision)
        else:
            data = self.get_data_from_map_file(input_file_1, input_file_2)
            
        data = shorten_list(data)
        write_into_excel(data, outputfolder, self.name_file_log)

        operations = {
            'excel': "Data Note",
            'zip': "Ref Drawing",
            'nx': "PDF CAD"
        }

        for op in search_type:
            label = operations[op]
            self.preparing(teamcenter_window,op)
            download_type(label)

            if self.stop_flag:
                self.done_label.config(text="Stopped!", fg="red")
            else:
                self.progress_label.config(text=f"Download {label} has finished!")
            self.download_button.config(state="normal")

        self.done_label.config(text="Done", fg="green")

        self.end_time = datetime.now().replace(microsecond=0)
        total_time = self.end_time - self.start_time

        print("----------------------------------------------------------")
        print(f"Start at: {self.start_time}")
        print(f"End at: {self.end_time}")
        print(f"Total Running Time: {total_time}")

        messagebox.showinfo("Completion", "The program has finished!")
        
    def main_function(self, input_file_1, input_file_2, outputfolder, coliteam, colrevision, colnamefolder, operation_type):
        if self.stop_flag: return
        app_path = "javaw.exe"
        self.app_nx = None
        self.window_nx = None
        self.export_window = None
        while True:
            if self.stop_flag: return
            try:
                app_teamcenter = Application(backend="uia").connect(path=app_path, timeout= 0.5)
                window_teamcenter = app_teamcenter.window(found_index=0)
                if window_teamcenter.exists():
                    break
            except:
                self.done_label.config(text="Please Login Teamcenter", fg="red")
                self.download_button.config(state="normal")
                return

        window_teamcenter.set_focus()
        window_teamcenter.maximize()
        
        self.main(app_teamcenter, window_teamcenter, input_file_1, input_file_2, outputfolder, operation_type, coliteam, colrevision, colnamefolder)

    def download(self):
        self.start_time = datetime.now().replace(microsecond=0)
        
        self.download_button.config(state="disabled")
        input_file_1 = self.input_file_entry_1.get()
        input_file_2 = self.input_file_entry_2.get()
        output_folder = self.output_folder_entry.get()
        coliteam = self.coliteam_entry.get()
        colrevision = self.colrevision_entry.get()
        colnamefolder = self.colnamefolder_entry.get()
        input_var = self.input_file_var.get()

        def is_alpha(value):
            return re.match("^[A-Za-z]+$", value) is not None

        if not input_file_1 and input_var or not input_file_2 and not input_var or not output_folder:
            self.download_button.config(state=tk.NORMAL)
            messagebox.showerror("Error", "Please select both input files and output folder.")
            return

        if not (is_alpha(coliteam) and is_alpha(colrevision) and is_alpha(colnamefolder)) and input_var:
            self.download_button.config(state=tk.NORMAL)
            messagebox.showerror("Error", "Please enter alphabetic characters for the columns.")
            return
        # Check if input files exist
        if input_var:
            if not os.path.isfile(input_file_1):
                self.download_button.config(state=tk.NORMAL)
                messagebox.showerror("Error", "Input file does not exist.")
                return
        else:
            if not os.path.isfile(input_file_1) or not os.path.isfile(input_file_2):
                self.download_button.config(state=tk.NORMAL)
                messagebox.showerror("Error", "One or both MAP/Connector IF files do not exist.")
                return
        # Build operation types based on checkboxes
        op_types = []
        if self.data_note_var.get() == 1:
            op_types.append('excel')
        if self.ref_drawing_var.get() == 1:
            op_types.append('zip')
        if self.pdf_cad_var.get() == 1:
            op_types.append('nx')

        if not op_types:
            self.download_button.config(state=tk.NORMAL)
            messagebox.showerror("Error", "Please select file type to download.")
            return
        self.progress_label.config(text="Preparing...", fg="blue")
        self.done_label.config(text="Downloading...", fg="green")
        self.stop_flag = False
        
        try:
            thread = threading.Thread(target=self.main_function, args=(input_file_1, input_file_2, output_folder, coliteam, colrevision, colnamefolder, op_types))
            thread.start()
        except:
            self.done_label.config(text="Error, Please try again", fg="red")
            self.download_button.config(state=tk.NORMAL)
    
    def stop_task(self):
        self.download_button.config(state="normal")
        self.progress_label.config(text="Stopped", fg="red")
        self.done_label.config(text="", fg="green")
        self.stop_flag = True

    def on_closing(self):
        self.save_data()
        self.root.destroy()

    def run(self):
        self.root.mainloop()    

if __name__ == "__main__":
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    if not os.path.exists("log"):
        os.makedirs("log")
    sys.stdout = Logger(os.path.join("log", f"log_{now}.txt"))
    sys.stderr = sys.stdout  # Redirect stderr to log as well
    # Hide the console window if running as a script
    try:
        ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)
    except Exception:
        pass
    app = TeamcenterDownloader()
    app.run()
